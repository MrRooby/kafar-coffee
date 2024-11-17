import discord
from discord.ext import commands, tasks
import os
from dotenv import load_dotenv
from datetime import datetime, timedelta
import xlsxwriter
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from database import *
from ganttchart import *

#TODO: - interakcja ma timeout 3min. Trzeba się zabezpieczyć przed tym

# Global variables
user_dyspo = {}

user_messages = {}

THICK_RIGHT_BORDER = Border(left=Side(style='thin'), 
                    right=Side(style='thick'), 
                    top=Side(style='none'), 
                    bottom=Side(style='none'))

THICK_LEFT_BORDER = Border(left=Side(style='thick'),
                    right=Side(style='thin'), 
                    top=Side(style='none'), 
                    bottom=Side(style='none'))

THICK_RIGHT_LEFT_BOTTOM_BORDER = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='none'), 
                    bottom=Side(style='thick'))


FILL_COLOURS = {
  "Monday": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),  # Desaturated Red
  "Tuesday": PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),  # Desaturated Orange
  "Wednesday": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),  # Desaturated Yellow
  "Thursday": PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid"),  # Desaturated Green
  "Friday": PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid"),  # Desaturated Light Blue
  "Saturday": PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid"),  # Desaturated Blue
  "Sunday": PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")   # Desaturated Purple
}

#for some reason PatternFill does not work for the make_spredsheet function so it has to be defined again
FILL_COLOURS_INIT = {
    "Monday": "#FF9999",  # Desaturated Red
    "Tuesday": "#FFCC99",  # Desaturated Orange
    "Wednesday": "#FFFF99",  # Desaturated Yellow
    "Thursday": "#99FF99",  # Desaturated Green
    "Friday": "#99CCFF",  # Desaturated Light Blue
    "Saturday": "#9999FF",  # Desaturated Blue
    "Sunday": "#CC99FF"   # Desaturated Purple
}


# Function to load the token from .env file
def load_token():
  load_dotenv()
  return os.getenv('TOKEN')


# Function to generate the dates for the current dyspo
def get_next_week_dates():
  today = datetime.today()
  start_of_next_week = today + timedelta(days=(7 - today.weekday()))
  end_of_next_week = start_of_next_week + timedelta(days=6)
  return start_of_next_week, end_of_next_week


# Function to initialize the bot with its permissions
def initialize_bot():
  intents = discord.Intents.default()
  intents.messages = True
  intents.guilds = True
  intents.message_content = True
  intents.members = True

  bot = commands.Bot(command_prefix='!', intents=intents)
  return bot

kafarDB = DATABASE()

def current_workbook_name():
  start_of_next_week, end_of_next_week = get_next_week_dates()
  return f"dyspozycje/dyspo[{start_of_next_week.strftime('%d')}-{end_of_next_week.strftime('%d.%m')}].xlsx"

# Function to create a new spreadsheet for the next week
def init_spredsheet():
  global current_workbook, current_worksheet, FILL_COLOURS_INIT
  file_name = current_workbook_name()
  current_workbook = xlsxwriter.Workbook(file_name)
  current_worksheet = current_workbook.add_worksheet()
  days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
  for row, day in enumerate(days_of_week):
    current_worksheet.write(row + 1, 0, day)
    cell_format = current_workbook.add_format({'bg_color': FILL_COLOURS_INIT[day]})
    current_worksheet.set_row(row + 1, None, cell_format)  # Apply the fill color to the row

  current_workbook.close()


def update_spredsheet(user, dyspo):
  global num_of_dyspo_from_users
  global current_worksheet, current_workbook
  global THICK_RIGHT_BORDER
  global FILL_COLOURS

  if not os.path.exists(current_workbook_name()):
    init_spredsheet()
  current_workbook = load_workbook(current_workbook_name())
  current_worksheet = current_workbook.active
  
  user_col = None
  for col in range(2, current_worksheet.max_column + 1):
    if current_worksheet.cell(row = 1, column = col).value == user:
      user_col = col
      break

  row = 1
  
  if user_col is None:
    user_col = current_worksheet.max_column + 1
    current_worksheet.cell(row = row, column = user_col).border = THICK_RIGHT_LEFT_BOTTOM_BORDER
    current_worksheet.cell(row = row, column = user_col + 1).border = THICK_RIGHT_LEFT_BOTTOM_BORDER

    current_worksheet.merge_cells(start_row = row, start_column = user_col, end_row = row, end_column = user_col + 1)    
    current_worksheet.cell(row = row, column = user_col).value = user

  row += 1

  
  # Define the correct order of days
  days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

  # Sort the dyspo dictionary according to the correct order of days
  sorted_dyspo = {day: dyspo[day] for day in days_order if day in dyspo}


  for day, hours in sorted_dyspo.items():
    if len(hours) == 2:
      current_worksheet.cell(row = row, column = user_col, value=hours[0])
      current_worksheet.cell(row = row, column = user_col + 1, value=hours[1])
      current_worksheet.cell(row = row, column = user_col).border = THICK_LEFT_BORDER
      current_worksheet.cell(row = row, column = user_col + 1).border = THICK_RIGHT_BORDER
    else:
      current_worksheet.cell(row = row, column = user_col, value=hours[0])
      current_worksheet.cell(row = row, column = user_col + 1, value=hours[0])
      current_worksheet.cell(row = row, column = user_col).border = THICK_LEFT_BORDER
      current_worksheet.cell(row = row, column = user_col + 1).border = THICK_RIGHT_BORDER

    current_worksheet.cell(row=row, column=user_col).fill = FILL_COLOURS[day]
    current_worksheet.cell(row=row, column=user_col + 1).fill = FILL_COLOURS[day]

    row += 1

  # create_gantt_chart(current_workbook, current_worksheet)

  current_workbook.save(current_workbook_name())


def add_to_spredsheet(user_id, day, hours):
  global num_of_dyspo_from_users

  if(user_id not in user_dyspo):
    user_dyspo[user_id] = {}

  # Handle special cases for "DO + 30min" and "OD + 30min"
  adjusted_hours = []
  do_30min_selected = False
  od_30min_selected = False

  for hour in hours:
    if hour == "DO + 30min":
      do_30min_selected = True
    elif hour == "OD + 30min":
      od_30min_selected = True
    else:
      adjusted_hours.append(hour)

  if do_30min_selected and adjusted_hours:
    adjusted_hours[0] = adjust_time(adjusted_hours[0], 30)
  if od_30min_selected and len(adjusted_hours) > 1:
    adjusted_hours[1] = adjust_time(adjusted_hours[1], 30)

  user_dyspo[user_id][day] = adjusted_hours


def adjust_time(time_str, minutes):
  hour, minute = map(int, time_str.split(":"))
  minute += minutes
  if minute >= 60:
    hour += 1
    minute -= 60
  return f"{hour:02d}:{minute:02d}"


async def notify_users(users):
  kafarDB.cursor.execute('SELECT USER_ID FROM NOTIFIED_USERS')
  notified_users = kafarDB.cursor.fetchall()
  for user_id in notified_users:
    user = await bot.fetch_user(user_id[0])
    await user.send(f"{users.name} wypełnił dyspo.")


# A class for a select menu for a day of the week
class WorkHoursSelect(discord.ui.Select):
  def __init__(self, placeholder):
    global num_of_dyspo_from_users
    options =  [discord.SelectOption(label="dowolnie")] + [discord.SelectOption(label="out")] + [
      discord.SelectOption(label=f"{hour}:00")
      for hour in range(6, 20)
    ] + [discord.SelectOption(label="OD + 30min")] + [discord.SelectOption(label="DO + 30min")]
    super().__init__(placeholder=placeholder, min_values=1, max_values=4, options=options)
    

  async def callback(self, interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    # Ensure "DO + 30min" and "OD + 30min" are not selected on their own
    if ("DO + 30min" in self.values or "OD + 30min" in self.values) and len(self.values) == 1:
      await interaction.followup.send("Options 'DO + 30min' and 'OD + 30min' must be selected with at least one other option.", ephemeral=True)
      return
    
    # Ensure "DO + 30min" and "OD + 30min" are not selected with "out" or "dowolnie"
    if ("DO + 30min" in self.values or "OD + 30min" in self.values) and ("out" in self.values or "dowolnie" in self.values):
      await interaction.followup.send("Options 'DO + 30min' and 'OD + 30min' cannot be selected with 'out' or 'dowolnie'.", ephemeral=True)
      return

    add_to_spredsheet(interaction.user.name, self.placeholder, self.values)


# 2 WorkHoursView classes due to discord limitation of 4 select options per view
class WorkHoursView1(discord.ui.View):
  def __init__(self):
    super().__init__()
    self.add_item(WorkHoursSelect("Monday"))
    self.add_item(WorkHoursSelect("Tuesday"))
    self.add_item(WorkHoursSelect("Wednesday"))
    self.add_item(WorkHoursSelect("Thursday"))
    self.timeout = 1800  # 30 minutes

class WorkHoursView2(discord.ui.View):
  def __init__(self):
    super().__init__()
    self.add_item(WorkHoursSelect("Friday"))
    self.add_item(WorkHoursSelect("Saturday"))
    self.add_item(WorkHoursSelect("Sunday"))
    self.timeout = 1800  # 30 minutes
    self.add_item(ConfirmButton())


class ConfirmButton(discord.ui.Button):
  def __init__(self):
    super().__init__(label="Confirm", style=discord.ButtonStyle.primary)
    self.timeout = 1800  # 30 minutes

  async def callback(self, interaction: discord.Interaction):
    user_id = interaction.user.name
    if(len(user_dyspo[user_id]) < 7):
      await interaction.response.send_message("Wypełnij wszystkie dni tygodnia", ephemeral=True)
      return
    update_spredsheet(interaction.user.name, user_dyspo[interaction.user.name])
    start_of_next_week, end_of_next_week = get_next_week_dates()
    kafarDB.add_dyspo_to_database(interaction.user.id, user_dyspo[interaction.user.name], start_of_next_week, end_of_next_week)
    confirmation_message = "Dyspo przesłane!\n\nWybrane godziny:\n"
    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for day in days_order:
      if day in user_dyspo[user_id]:
        hours = user_dyspo[user_id][day]
        confirmation_message += f"{day}: {'-'.join(hours)}\n"
    print(interaction.user.id)
    await interaction.response.send_message(confirmation_message, ephemeral=True)

    await notify_users(interaction.user)

    await interaction.message.delete()

    if user_id in user_messages:
      for message in user_messages[user_id]:
          try:
            await message.delete()
          except discord.errors.NotFound:
            print(f"Message {message.id} not found, skipping deletion.")
      del user_messages[user_id]
      

async def send_select_menus(user):
  start_date, end_date = get_next_week_dates()
  view1 = WorkHoursView1()
  view2 = WorkHoursView2()
  date = "[Dyspo " + start_date.strftime('%d') + "-" + end_date.strftime('%d.%m') + "]\n"
  instructions = "W celu wybrania połowy godziny dodaj do wybranej godziny 'DO + 30min' lub 'OD + 30min'.\n"
  message1 = await user.send(date + instructions, view=view1)
  message2 = await user.send(view=view2)
  
  # Store the messages to delete later
  user_messages[user.name] = [message1, message2]




bot = initialize_bot()

@bot.event
async def on_ready():
  print(f"Logged in as {bot.user}")
  send_dyspo.start()


@bot.command()
async def dyspo(ctx):
  await send_select_menus(ctx.author)


@bot.command()
async def excel(ctx):
  if os.path.exists(current_workbook_name()):
    await ctx.send(file=discord.File(current_workbook_name()))
  else:
    await ctx.send("No dyspo sheet found")


@bot.command()
async def notifon(ctx):
  kafarDB.cursor.execute('SELECT * FROM NOTIFIED_USERS WHERE USER_ID = %s', (ctx.author.id,))
  result = kafarDB.cursor.fetchone()

  if result is None:
    kafarDB.cursor.execute('INSERT INTO NOTIFIED_USERS (USER_ID) VALUES (%s)', (ctx.author.id,))
    kafarDB.connection.commit()
    await ctx.send(":green_circle: Włączono powiadomienia!")
  else:
    await ctx.send("Powiadomienia już są włączone!")



@bot.command()
async def notifoff(ctx):
  kafarDB.cursor.execute('SELECT * FROM NOTIFIED_USERS WHERE USER_ID = %s', (ctx.author.id,))
  result = kafarDB.cursor.fetchone()

  if result is None:
    await ctx.send("Powiadomienia już są wyłączone!")
  else:
    kafarDB.cursor.execute("DELETE FROM NOTIFIED_USERS WHERE USER_ID = %s", (ctx.author.id,))
    kafarDB.connection.commit()
    await ctx.send(":red_circle: Wyłączono powiadomienia!")


@bot.command()
async def wykres(ctx):
  start_of_next_week, end_of_next_week = get_next_week_dates()
  switcher = {
    1: 'MON',
    2: 'TUE',
    3: 'WED',
    4: 'THU',
    5: 'FRI',
    6: 'SAT',
    7: 'SUN'
  }

  for day in range(1, 8):
    create_gantt_chart(start_of_next_week, end_of_next_week, day, kafarDB.connection, kafarDB.cursor)
    day_name = (start_of_next_week + timedelta(days=day-1)).strftime('%A')
    await ctx.send(file=discord.File(f"charts/{switcher[day]}.png"))


@tasks.loop(hours=1) # The message will be delivered every wednesday form 16 to 16:59 depending when was the program started
async def send_dyspo():
  now = datetime.now()
  if (now.weekday() == 2 and now.hour == 16):  # Check if it's Wednesday at 4 PM
    for guild in bot.guilds:
      for member in guild.members:
        start_of_next_week, end_of_next_week = get_next_week_dates()
        if not kafarDB.is_form_sent_record_exists(member.id, start_of_next_week, end_of_next_week) and not member.bot and discord.utils.get(member.roles, name="beboki") and kafarDB.dyspo_record_exists(member.id, start_of_next_week, end_of_next_week) == False:  # Skip bot accounts and check for role "beboki"
          await send_select_menus(member)
          kafarDB.add_form_sent_record(member.id, start_of_next_week, end_of_next_week)


@tasks.loop(hours=1)
async def create_spredsheet():
  global current_worksheet, current_workbook, num_of_dyspo_from_users
  now = datetime.now()
  if (now.weekday() == 0 and now.hour == 0):
    current_workbook, current_worksheet = init_spredsheet()
    num_of_dyspo_from_users = 0
    print("New dyspo sheet created")


bot.run(load_token())