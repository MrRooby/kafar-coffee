################################################################################################################
################################ Formatting for excel spreadsheet ##############################################
################################################################################################################

from openpyxl.styles import PatternFill, Border, Side

THICK_RIGHT_BORDER = Border(left   = Side(style='thin'), 
                            right  = Side(style='thick'), 
                            top    = Side(style='none'), 
                            bottom = Side(style='none'))

THICK_LEFT_BORDER = Border(left   = Side(style='thick'),
                           right  = Side(style='thin'), 
                           top    = Side(style='none'), 
                           bottom = Side(style='none'))

THICK_RIGHT_LEFT_BOTTOM_BORDER = Border(left   = Side(style='thick'), 
                                        right  = Side(style='thick'), 
                                        top    = Side(style='none'), 
                                        bottom = Side(style='thick'))

FILL_COLOURS = {
  "Monday":    PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),  # Desaturated Red
  "Tuesday":   PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),  # Desaturated Orange
  "Wednesday": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),  # Desaturated Yellow
  "Thursday":  PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid"),  # Desaturated Green
  "Friday":    PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid"),  # Desaturated Light Blue
  "Saturday":  PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid"),  # Desaturated Blue
  "Sunday":    PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")   # Desaturated Purple
}

#for some reason PatternFill does not work for the make_spredsheet function so it has to be defined again
FILL_COLOURS_INIT = {
    "Monday":    "#FF9999", # Desaturated Red
    "Tuesday":   "#FFCC99", # Desaturated Orange
    "Wednesday": "#FFFF99", # Desaturated Yellow
    "Thursday":  "#99FF99", # Desaturated Green
    "Friday":    "#99CCFF", # Desaturated Light Blue
    "Saturday":  "#9999FF", # Desaturated Blue
    "Sunday":    "#CC99FF"  # Desaturated Purple
}

DAYS_ORDER = ["Monday",
              "Tuesday",
              "Wednesday",
              "Thursday",
              "Friday",
              "Saturday",
              "Sunday"]
################################################################################################################