import xlsxwriter
from openpyxl import load_workbook
from datetime import datetime, timedelta
from constants import *
import os


class Spredsheet:
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        pass


    def __get_next_week_dates(self):
        """Generates the starting and ending dates for the current week.

        Returns:
            tuple: A tuple containing:
            - datetime: The date of the first day of the week.
            - datetime: The date of the last day of the week.
        """
        today = datetime.today()
        start_of_next_week = today + timedelta(days=(7 - today.weekday()))
        end_of_next_week = start_of_next_week + timedelta(days=6)
        return start_of_next_week, end_of_next_week


    def __current_workbook_name(self):
        """Generates a name for the Excel file with the current week's dates.

        Returns:
          str: Name for the Excel file.
        """
        start_of_next_week, end_of_next_week = self.__get_next_week_dates()
        return f"dyspozycje/dyspo[{start_of_next_week.strftime('%d')}-{end_of_next_week.strftime('%d.%m')}].xlsx"


    def init_spredsheet(self):
        '''Initializes a new spreadsheet for current week.

        Fills days of the week and applies initial fill colors.
        This function creates a new Excel workbook and worksheet using the xlsxwriter library.
        It writes the days of the week in the first column and applies a background color to each row
        based on the predefined `FILL_COLOURS_INIT` dictionary.
        
        Globals:
            current_workbook (xlsxwriter.Workbook): The current workbook being created.
            current_worksheet (xlsxwriter.Workbook.worksheet_class): The current worksheet being created.
            FILL_COLOURS_INIT (dict): A dictionary mapping days of the week to their respective fill colors.
        
        Raises:
            KeyError: If a day of the week is not found in the `FILL_COLOURS_INIT` dictionary.
        '''
        file_name = self.current_workbook_name()
        self.workbook = xlsxwriter.Workbook(file_name)
        self.worksheet = self.workbook.add_worksheet()

        days_of_week = ["Monday", 
                        "Tuesday",
                        "Wednesday",
                        "Thursday",
                        "Friday",
                        "Saturday",
                        "Sunday"]

        for row, day in enumerate(days_of_week):
            self.worksheet.write(row + 1, 0, day)
            cell_format = self.workbook.add_format({'bg_color': FILL_COLOURS_INIT[day]})
            self.worksheet.set_row(row + 1, None, cell_format)  # Apply the fill color to the row

        self.workbook.close()


    def __add_user_columns(user, worksheet, user_col):
        # Check if user already exists in the spredsheet
        for col in range(2, worksheet.max_column + 1):
            if worksheet.cell(row = 1, column = col).value == user:
                user_col = col
            break

        row = 1
        
        # If user doesn't exist in the spredsheet create space for them
        if user_col is None:
            user_col = worksheet.max_column + 1
            worksheet.cell(row = row, column = user_col).border = THICK_RIGHT_LEFT_BOTTOM_BORDER
            worksheet.cell(row = row, column = user_col + 1).border = THICK_RIGHT_LEFT_BOTTOM_BORDER

            worksheet.merge_cells(start_row = row, 
                                  start_column = user_col, 
                                  end_row = row, 
                                  end_column = user_col + 1)    
            worksheet.cell(row = row, column = user_col).value = user

        row += 1


    def __fill_user_columns(user_col, sorted_dyspo: dict, worksheet):
        for day, hours in sorted_dyspo.items():
            if len(hours) == 2:
                worksheet.cell(row = row, column = user_col,     value=hours[0])
                worksheet.cell(row = row, column = user_col + 1, value=hours[1])
                worksheet.cell(row = row, column = user_col)    .border = THICK_LEFT_BORDER
                worksheet.cell(row = row, column = user_col + 1).border = THICK_RIGHT_BORDER
            else:
                worksheet.cell(row = row, column = user_col,     value=hours[0])
                worksheet.cell(row = row, column = user_col + 1, value=hours[0])
                worksheet.cell(row = row, column = user_col)    .border = THICK_LEFT_BORDER
                worksheet.cell(row = row, column = user_col + 1).border = THICK_RIGHT_BORDER

            worksheet.cell(row=row, column=user_col)    .fill = FILL_COLOURS[day]
            worksheet.cell(row=row, column=user_col + 1).fill = FILL_COLOURS[day]

            row += 1


    def update_spredsheet(self, 
                          user, 
                          dyspo, 
                          worksheet, workbook):
        if not os.path.exists(self.__current_workbook_name()):
            self.__init_spredsheet()

        workbook = load_workbook(self.__current_workbook_name())
        worksheet = workbook.active
        
        # Assume that user is not present in the spredsheet
        user_col = None
        self.__add_user_columns(user, worksheet, user_col)
        
        # Sort the dyspo dictionary according to the correct order of days
        # This is done to sort data from user who might have entered dyspo 
        # in random order
        sorted_dyspo = {day: dyspo[day] for day in DAYS_ORDER if day in dyspo}

        self.__fill_user_columns(user_col, sorted_dyspo, worksheet)

        workbook.save(self.current_workbook_name())


    def add_dyspo(self, user_id, day, hours, user_dyspo):
        if(user_id not in user_dyspo):
            user_dyspo[user_id] = {}

        user_dyspo[user_id][day] = hours